import datetime
import io
import json
from decimal import Decimal

from django.core.files.base import ContentFile
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from apps.postulaciones.models import Postulacion
from apps.postulaciones.signals import resultado_adjudicacion

from . import charts, charts_matplotlib, llm_client, llm_tools, selectors
from .models import Conversacion, MensajeChat, ResumenIA


def procesar_formularios_socioeconomicos(*, convocatoria):
    """CU23 — Calcula el puntaje socioeconómico para postulaciones APROBADAS.

    Usa Pandas para normalizar y ponderar los factores socioeconómicos.
    Actualiza cada Postulacion con su puntaje y la marca como PROCESADA.

    Args:
        convocatoria: Instancia de Convocatoria.

    Returns:
        Cantidad de postulaciones procesadas.
    """
    import pandas as pd

    postulaciones = list(
        Postulacion.objects.filter(
            convocatoria=convocatoria,
            estado=Postulacion.Estado.APROBADA,
        ).select_related("formulario")
    )

    if not postulaciones:
        return 0

    data = [
        {
            "id": p.pk,
            "ingreso": float(p.formulario.ingreso_mensual_familiar),
            "familiares": p.formulario.cantidad_familiares,
            "desempleado": p.formulario.situacion_laboral == "DESEMPLEADO",
            "no_propietario": p.formulario.situacion_habitacional != "PROPIETARIO",
            "sin_beca_previa": not p.formulario.tiene_beca_previa,
        }
        for p in postulaciones
    ]

    df = pd.DataFrame(data)

    ingreso_max = df["ingreso"].max()
    familiares_max = df["familiares"].max()

    df["ingreso_norm"] = df["ingreso"] / ingreso_max if ingreso_max > 0 else 0.0
    df["familiares_norm"] = df["familiares"] / familiares_max if familiares_max > 0 else 0.0

    df["puntaje"] = (
        (1 - df["ingreso_norm"]) * 40
        + df["desempleado"].astype(int) * 20
        + df["familiares_norm"] * 20
        + df["no_propietario"].astype(int) * 10
        + df["sin_beca_previa"].astype(int) * 10
    ).round(2)

    for _, row in df.iterrows():
        Postulacion.objects.filter(pk=int(row["id"])).update(
            puntaje_socioeconomico=Decimal(str(row["puntaje"])),
            estado=Postulacion.Estado.PROCESADA,
        )

    return len(postulaciones)


def generar_ranking(*, convocatoria, cupo, cupo_espera=None):
    """CU24 — Ordena postulaciones PROCESADAS y asigna resultados finales.

    Args:
        convocatoria: Instancia de Convocatoria.
        cupo: Número de postulaciones a adjudicar.
        cupo_espera: Número de postulaciones en lista de espera (default = cupo).

    Returns:
        Lista de Postulacion con estado actualizado, ordenadas por puntaje.
    """
    if cupo_espera is None:
        cupo_espera = cupo

    postulaciones = list(
        Postulacion.objects.filter(
            convocatoria=convocatoria,
            estado=Postulacion.Estado.PROCESADA,
        )
        .select_related("estudiante", "beca")
        .order_by("-puntaje_socioeconomico", "fecha_envio")
    )

    for i, p in enumerate(postulaciones):
        if i < cupo:
            p.estado = Postulacion.Estado.ADJUDICADA
        elif i < cupo + cupo_espera:
            p.estado = Postulacion.Estado.LISTA_ESPERA
        else:
            p.estado = Postulacion.Estado.NO_ADJUDICADA

        p.save(update_fields=["estado"])
        resultado_adjudicacion.send(sender=Postulacion, postulacion=p)

    return postulaciones


def exportar_ranking_excel(*, convocatoria):
    """CU25 — Genera un archivo Excel con el ranking de la convocatoria.

    Args:
        convocatoria: Instancia de Convocatoria.

    Returns:
        Bytes del archivo .xlsx.
    """
    postulaciones = (
        Postulacion.objects.filter(
            convocatoria=convocatoria,
            estado__in=[
                Postulacion.Estado.ADJUDICADA,
                Postulacion.Estado.LISTA_ESPERA,
                Postulacion.Estado.NO_ADJUDICADA,
            ],
        )
        .select_related("estudiante", "beca")
        .order_by("-puntaje_socioeconomico", "fecha_envio")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Ranking"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    headers = ["Pos.", "Apellido", "Nombre", "Email", "Beca", "Puntaje", "Resultado"]
    col_widths = [6, 20, 20, 30, 25, 12, 20]

    for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = width

    estado_colores = {
        Postulacion.Estado.ADJUDICADA: "C6EFCE",
        Postulacion.Estado.LISTA_ESPERA: "FFEB9C",
        Postulacion.Estado.NO_ADJUDICADA: "FFC7CE",
    }

    for pos, p in enumerate(postulaciones, start=1):
        fila = [
            pos,
            p.estudiante.last_name,
            p.estudiante.first_name,
            p.estudiante.email,
            p.beca.nombre,
            float(p.puntaje_socioeconomico or 0),
            p.get_estado_display(),
        ]
        color = estado_colores.get(p.estado, "FFFFFF")
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col, valor in enumerate(fila, start=1):
            cell = ws.cell(row=pos + 1, column=col, value=valor)
            cell.fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def exportar_reporte_pdf(*, convocatoria, request=None):
    """CU26 — Genera un PDF con el reporte de la convocatoria.

    Args:
        convocatoria: Instancia de Convocatoria.
        request: HttpRequest opcional (para URLs absolutas en el PDF).

    Returns:
        Bytes del archivo .pdf.
    """
    from weasyprint import HTML

    postulaciones = list(
        Postulacion.objects.filter(
            convocatoria=convocatoria,
            estado__in=[
                Postulacion.Estado.ADJUDICADA,
                Postulacion.Estado.LISTA_ESPERA,
                Postulacion.Estado.NO_ADJUDICADA,
            ],
        )
        .select_related("estudiante", "beca")
        .order_by("-puntaje_socioeconomico", "fecha_envio")
    )

    total = len(postulaciones)
    adjudicadas = sum(1 for p in postulaciones if p.estado == Postulacion.Estado.ADJUDICADA)
    espera = sum(1 for p in postulaciones if p.estado == Postulacion.Estado.LISTA_ESPERA)
    no_adjudicadas = total - adjudicadas - espera

    contexto = {
        "convocatoria": convocatoria,
        "postulaciones": postulaciones,
        "total": total,
        "adjudicadas": adjudicadas,
        "espera": espera,
        "no_adjudicadas": no_adjudicadas,
    }

    html_str = render_to_string("reportes/reporte_pdf.html", contexto)
    pdf_bytes = HTML(string=html_str).write_pdf()
    return pdf_bytes


def construir_contexto_dashboard(*, convocatoria=None, fecha_desde=None, fecha_hasta=None):
    """CU26 — Orquesta selectors.py + charts.py para el dashboard de KPIs.

    Args:
        convocatoria: Instancia de Convocatoria para filtrar, o None para todas.
        fecha_desde: date mínima de fecha_creacion a incluir, o None.
        fecha_hasta: date máxima de fecha_creacion a incluir, o None.

    Returns:
        Dict listo para renderizar en dashboard.html / _dashboard_kpis.html.
    """
    filtro = {"convocatoria": convocatoria, "fecha_desde": fecha_desde, "fecha_hasta": fecha_hasta}

    demanda_beca = selectors.postulaciones_por_beca(**filtro)
    tasa_adj_conv = selectors.tasa_adjudicacion_por_convocatoria()
    espera_conv = selectors.tamano_lista_espera_por_convocatoria()

    embudo = selectors.embudo_estados(**filtro)
    rechazos = selectors.desglose_rechazos(**filtro)
    tiempos_etapa = selectors.tiempos_promedio_por_etapa(convocatoria=convocatoria)

    validacion_doc = selectors.validacion_por_tipo_documento()
    mayor_rechazo = selectors.documento_mayor_rechazo()

    ingreso = selectors.distribucion_ingreso_familiar(convocatoria=convocatoria)
    puntaje = selectors.distribucion_puntaje_socioeconomico(convocatoria=convocatoria)
    indicadores = selectors.indicadores_generales(convocatoria=convocatoria)

    punto_corte = selectors.punto_corte_por_convocatoria()

    carreras = selectors.postulantes_por_carrera(convocatoria=convocatoria)
    tasa_carrera = selectors.tasa_adjudicacion_por_carrera()
    anio_ingreso = selectors.postulantes_por_anio_ingreso(convocatoria=convocatoria)

    entrega = selectors.tasa_entrega_notificaciones(
        fecha_desde=fecha_desde, fecha_hasta=fecha_hasta
    )
    latencia = selectors.latencia_envio_notificaciones()

    if convocatoria is not None:
        comparacion = selectors.comparacion_puntaje_por_resultado(convocatoria=convocatoria)
        grafico_comparacion = charts.fig_boxplot(
            datos=comparacion, titulo="Puntaje socioeconómico por resultado"
        )
    else:
        grafico_comparacion = charts.vacio(
            "Seleccioná una convocatoria específica para comparar puntajes por resultado."
        )

    return {
        "grafico_demanda_beca": charts.fig_barras(**demanda_beca, titulo="Postulaciones por beca"),
        "grafico_tasa_adjudicacion_convocatoria": charts.fig_barras(
            **tasa_adj_conv,
            titulo="Tasa de adjudicación por convocatoria",
            horizontal=True,
            sufijo="%",
        ),
        "grafico_lista_espera": charts.fig_barras(
            **espera_conv, titulo="Tamaño de lista de espera por convocatoria"
        ),
        "grafico_embudo": charts.fig_funnel(**embudo, titulo="Embudo de postulaciones"),
        "grafico_rechazos": charts.fig_donut(**rechazos, titulo="Motivos de rechazo"),
        "grafico_tiempos_etapa": charts.fig_barras(
            **tiempos_etapa, titulo="Tiempo promedio entre etapas", horizontal=True, sufijo=" días"
        ),
        "grafico_validacion_documental": charts.fig_barras_apiladas(
            etiquetas=validacion_doc["etiquetas"],
            series={
                "Aprobado": validacion_doc["aprobado"],
                "Rechazado": validacion_doc["rechazado"],
                "Pendiente": validacion_doc["pendiente"],
            },
            titulo="Validación documental por tipo",
        ),
        "documento_mayor_rechazo": mayor_rechazo,
        "grafico_ingreso_familiar": charts.fig_histograma(
            valores=ingreso["valores"],
            titulo="Distribución de ingreso familiar",
            eje_x="Ingreso mensual familiar (ARS)",
        ),
        "grafico_puntaje": charts.fig_histograma(
            valores=puntaje["valores"],
            titulo="Distribución de puntaje socioeconómico",
            eje_x="Puntaje",
        ),
        "grafico_comparacion_puntaje": grafico_comparacion,
        "graficos_distribucion_choices": {
            campo: charts.fig_donut(
                **selectors.distribucion_choices(campo, convocatoria=convocatoria), titulo=titulo
            )
            for campo, titulo in [
                ("situacion_laboral", "Situación laboral"),
                ("situacion_habitacional", "Situación habitacional"),
                ("dependencia_economica", "Dependencia económica"),
                ("tipo_tenencia_vivienda", "Tenencia de vivienda"),
            ]
        },
        "indicadores_generales": indicadores,
        "punto_corte_por_convocatoria": list(
            zip(punto_corte["etiquetas"], punto_corte["valores"], strict=True)
        ),
        "grafico_carreras": charts.fig_barras(
            **carreras, titulo="Postulantes por carrera (top 10)", horizontal=True
        ),
        "grafico_tasa_carrera": charts.fig_barras(
            **tasa_carrera, titulo="Tasa de adjudicación por carrera", horizontal=True, sufijo="%"
        ),
        "grafico_anio_ingreso": charts.fig_barras(
            **anio_ingreso, titulo="Postulantes por año de ingreso"
        ),
        "grafico_entrega_notificaciones": charts.fig_donut(
            **entrega, titulo="Estado de envío de notificaciones"
        ),
        "latencia_notificaciones": latencia,
    }


def solicitar_resumen_ia(*, usuario, convocatoria=None, prompt_adicional=""):
    """CU26 — Encola la generación de un resumen narrativo de KPIs con el LLM local.

    Args:
        usuario: Usuario (Director) que solicita el resumen.
        convocatoria: Instancia de Convocatoria a resumir, o None para todas.
        prompt_adicional: Instrucción libre opcional del Director.

    Returns:
        La instancia de ResumenIA creada (estado PENDIENTE).
    """
    resumen = ResumenIA.objects.create(
        usuario=usuario, convocatoria=convocatoria, prompt_adicional=prompt_adicional
    )
    from .tasks import tarea_generar_resumen_ia

    tarea_generar_resumen_ia.delay(resumen.pk)
    return resumen


def _contexto_kpis_para_prompt(*, convocatoria):
    """Serializa los KPIs reales (selectors.py) en texto plano para el prompt del LLM."""
    lineas = [
        f"Embudo de postulaciones por estado: {selectors.embudo_estados(convocatoria=convocatoria)}",
        f"Demanda por beca: {selectors.postulaciones_por_beca(convocatoria=convocatoria)}",
        f"Motivos de rechazo: {selectors.desglose_rechazos(convocatoria=convocatoria)}",
        f"Indicadores socioeconómicos: {selectors.indicadores_generales(convocatoria=convocatoria)}",
    ]
    if convocatoria is not None:
        lineas.append(
            f"Puntaje socioeconómico por resultado final: "
            f"{selectors.comparacion_puntaje_por_resultado(convocatoria=convocatoria)}"
        )
    else:
        lineas.append(
            f"Tasa de adjudicación por convocatoria: {selectors.tasa_adjudicacion_por_convocatoria()}"
        )
    return "\n".join(lineas)


def generar_resumen_con_llm(*, resumen_pk):
    """Construye el prompt a partir de KPIs reales y llama al LLM local (Ollama).

    Llamada exclusivamente por `tasks.tarea_generar_resumen_ia` (cola "ia", procesada
    solo por el worker con acceso a Ollama).
    """
    resumen = ResumenIA.objects.select_related("convocatoria").get(pk=resumen_pk)
    resumen.estado = ResumenIA.Estado.PROCESANDO
    resumen.save(update_fields=["estado"])

    try:
        contexto = _contexto_kpis_para_prompt(convocatoria=resumen.convocatoria)
        mensajes = [
            {
                "role": "system",
                "content": (
                    "Eres un asistente que redacta resúmenes ejecutivos para el Director "
                    "de un sistema de becas universitarias. Basate ÚNICAMENTE en los datos "
                    "numéricos provistos a continuación; no inventes cifras que no estén en "
                    "el contexto. Responde en español, en un párrafo conciso."
                ),
            },
            {"role": "user", "content": f"{contexto}\n\n{resumen.prompt_adicional}".strip()},
        ]
        respuesta = llm_client.chat(mensajes)
        resumen.resultado = respuesta["message"]["content"]
        resumen.estado = ResumenIA.Estado.COMPLETADO
        resumen.fecha_completado = timezone.now()
        resumen.save(update_fields=["resultado", "estado", "fecha_completado"])
    except Exception as exc:
        resumen.estado = ResumenIA.Estado.ERROR
        resumen.error_detalle = str(exc)
        resumen.save(update_fields=["estado", "error_detalle"])
        raise


def marcar_resumenes_ia_vencidos(*, minutos=10):
    """Marca como ERROR los ResumenIA atascados en PENDIENTE/PROCESANDO (LLM no disponible).

    Returns:
        Cantidad de registros marcados como error.
    """
    limite = timezone.now() - datetime.timedelta(minutes=minutos)
    return ResumenIA.objects.filter(
        estado__in=[ResumenIA.Estado.PENDIENTE, ResumenIA.Estado.PROCESANDO],
        fecha_creacion__lt=limite,
    ).update(
        estado=ResumenIA.Estado.ERROR,
        error_detalle="El servicio de IA no está disponible en este momento. Intentá nuevamente más tarde.",
    )


def exportar_resumen_ia_pdf(*, resumen_pk):
    """Genera un PDF con el resumen narrativo de IA + gráficos estáticos (Matplotlib).

    Returns:
        Bytes del archivo .pdf.
    """
    from weasyprint import HTML

    resumen = ResumenIA.objects.select_related("convocatoria").get(pk=resumen_pk)

    embudo = selectors.embudo_estados(convocatoria=resumen.convocatoria)
    puntaje = selectors.distribucion_puntaje_socioeconomico(convocatoria=resumen.convocatoria)

    contexto = {
        "resumen": resumen,
        "grafico_embudo": charts_matplotlib.grafico_embudo_estatico(
            **embudo, titulo="Embudo de postulaciones"
        ),
        "grafico_puntaje": charts_matplotlib.grafico_histograma_estatico(
            valores=puntaje["valores"],
            titulo="Distribución de puntaje socioeconómico",
            eje_x="Puntaje",
        ),
    }
    html_str = render_to_string("reportes/resumen_ia_pdf.html", contexto)
    return HTML(string=html_str).write_pdf()


_COLUMNAS_POSTULANTES = [
    ("nombre", "Nombre"),
    ("email", "Email"),
    ("legajo", "Legajo"),
    ("carrera", "Carrera"),
    ("anio_ingreso", "Año ingreso"),
    ("cantidad_familiares", "Familiares"),
    ("convocatoria", "Convocatoria"),
    ("beca", "Beca"),
    ("estado_postulacion", "Estado"),
]


def generar_archivo_postulantes(*, filas, formato):
    """Genera un archivo descargable con una lista de postulantes ya filtrada.

    Args:
        filas: Lista de dicts, como las que devuelve `selectors.buscar_postulantes`.
        formato: "excel" o "pdf".

    Returns:
        Tupla (nombre_archivo, bytes).
    """
    marca = timezone.now().strftime("%Y%m%d%H%M%S")

    if formato == "pdf":
        from weasyprint import HTML

        html_str = render_to_string("reportes/postulantes_pdf.html", {"filas": filas})
        return f"reporte_postulantes_{marca}.pdf", HTML(string=html_str).write_pdf()

    wb = Workbook()
    ws = wb.active
    ws.title = "Postulantes"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    encabezados = [titulo for _, titulo in _COLUMNAS_POSTULANTES]
    for col, encabezado in enumerate(encabezados, start=1):
        cell = ws.cell(row=1, column=col, value=encabezado)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for fila_idx, fila in enumerate(filas, start=2):
        for col, (clave, _) in enumerate(_COLUMNAS_POSTULANTES, start=1):
            ws.cell(row=fila_idx, column=col, value=fila.get(clave))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return f"reporte_postulantes_{marca}.xlsx", buf.read()


_ROL_A_OLLAMA = {
    MensajeChat.Rol.USUARIO: "user",
    MensajeChat.Rol.ASISTENTE: "assistant",
    MensajeChat.Rol.SISTEMA: "system",
}

_PROMPT_SISTEMA_CHAT = (
    "Eres un asistente que ayuda al Director de un sistema de becas universitarias a "
    "consultar datos reales del sistema. Para responder preguntas sobre datos "
    "(postulaciones, becas, convocatorias, documentos, notificaciones), SIEMPRE usá las "
    "herramientas disponibles en vez de inventar cifras. Si necesitás el id de una "
    "convocatoria, llamá primero a listar_convocatorias. Si te piden una tabla, listado "
    "o reporte descargable, usá generar_reporte_postulantes con los filtros que puedas "
    "inferir — nunca intentes escribir la tabla vos mismo como texto, el archivo se "
    "genera aparte. El sistema NO registra la edad ni la fecha de nacimiento del "
    "estudiante: si te piden filtrar o reportar por edad, decilo explícitamente en vez "
    "de inventar un valor o usar otro campo como reemplazo. Respondé en español, de "
    "forma clara y concisa, citando los números reales que obtuviste de las herramientas."
)

MAX_ITERACIONES_TOOLS = 5


def crear_conversacion(*, usuario):
    """Crea una conversación de chat vacía para el usuario."""
    return Conversacion.objects.create(usuario=usuario)


def enviar_mensaje_chat(*, conversacion, contenido):
    """Guarda el mensaje del usuario y encola el procesamiento con el LLM local.

    Returns:
        El MensajeChat del usuario recién creado.
    """
    mensaje = MensajeChat.objects.create(
        conversacion=conversacion, rol=MensajeChat.Rol.USUARIO, contenido=contenido
    )
    if not conversacion.titulo:
        conversacion.titulo = contenido[:60]
    conversacion.save()  # actualiza fecha_actualizacion (auto_now) y, si corresponde, el título

    from .tasks import tarea_procesar_mensaje_chat

    tarea_procesar_mensaje_chat.delay(mensaje.pk)
    return mensaje


def procesar_mensaje_con_llm(*, mensaje_pk):
    """Genera la respuesta del asistente para el último mensaje de una conversación.

    Usa tool calling: si el modelo pide ejecutar una herramienta, se ejecuta vía
    `llm_tools.ejecutar_tool` (lista blanca) y el resultado se le devuelve al modelo
    antes de pedirle la respuesta final. Llamada exclusivamente por
    `tasks.tarea_procesar_mensaje_chat` (cola "ia").
    """
    mensaje_usuario = MensajeChat.objects.select_related("conversacion").get(pk=mensaje_pk)
    conversacion = mensaje_usuario.conversacion

    historial = conversacion.mensajes.order_by("fecha_creacion")
    mensajes_llm = [{"role": "system", "content": _PROMPT_SISTEMA_CHAT}]
    mensajes_llm += [{"role": _ROL_A_OLLAMA[m.rol], "content": m.contenido} for m in historial]

    tools_usadas = []
    archivo_generado = None
    try:
        contenido_final = (
            "No pude completar la consulta (demasiados pasos). Probá reformular la pregunta."
        )
        for _ in range(MAX_ITERACIONES_TOOLS):
            respuesta = llm_client.chat(mensajes_llm, tools=llm_tools.definiciones_tools())
            mensaje_llm = respuesta["message"]
            tool_calls = mensaje_llm.get("tool_calls") or []

            if not tool_calls:
                contenido_final = mensaje_llm.get("content", "")
                break

            mensajes_llm.append(
                {
                    "role": "assistant",
                    "content": mensaje_llm.get("content", ""),
                    "tool_calls": tool_calls,
                }
            )
            for llamada in tool_calls:
                nombre = llamada["function"]["name"]
                argumentos = llamada["function"].get("arguments") or {}
                resultado = llm_tools.ejecutar_tool(nombre, argumentos)
                archivo_bytes = resultado.pop("_archivo_bytes", None)
                archivo_nombre = resultado.pop("_archivo_nombre", None)
                if archivo_bytes:
                    archivo_generado = (archivo_nombre, archivo_bytes)
                tools_usadas.append({"tool": nombre, "argumentos": argumentos})
                mensajes_llm.append(
                    {"role": "tool", "name": nombre, "content": json.dumps(resultado, default=str)}
                )

        mensaje_asistente = MensajeChat.objects.create(
            conversacion=conversacion,
            rol=MensajeChat.Rol.ASISTENTE,
            contenido=contenido_final,
            tools_usadas=tools_usadas or None,
        )
        if archivo_generado is not None:
            archivo_nombre, archivo_bytes = archivo_generado
            mensaje_asistente.archivo.save(archivo_nombre, ContentFile(archivo_bytes), save=True)
        conversacion.save()  # toca fecha_actualizacion
    except Exception as exc:
        MensajeChat.objects.create(
            conversacion=conversacion,
            rol=MensajeChat.Rol.ASISTENTE,
            contenido=f"No pude responder: el servicio de IA no está disponible ({exc}).",
        )
        conversacion.save()
        raise


def marcar_chats_vencidos(*, minutos=10):
    """Inserta una respuesta de error en conversaciones cuyo último mensaje es del
    usuario y quedó sin contestar por más de `minutos` (LLM no disponible).

    Returns:
        Cantidad de conversaciones marcadas como vencidas.
    """
    limite = timezone.now() - datetime.timedelta(minutes=minutos)
    count = 0
    for conversacion in Conversacion.objects.all():
        ultimo = conversacion.mensajes.order_by("-fecha_creacion").first()
        if (
            ultimo is not None
            and ultimo.rol == MensajeChat.Rol.USUARIO
            and ultimo.fecha_creacion < limite
        ):
            MensajeChat.objects.create(
                conversacion=conversacion,
                rol=MensajeChat.Rol.ASISTENTE,
                contenido="El servicio de IA no está disponible en este momento. Intentá nuevamente más tarde.",
            )
            conversacion.save()
            count += 1
    return count
